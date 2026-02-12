import csv
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging

logger = logging.getLogger(__name__)


def export_user_actions_to_csv(actions, filename=None):
    """
    actions — список словарей:
    [
      {
        'user_id': ...,
        'telegram_id': ...,
        'username': ...,
        'action': ...,
        'details': ...,
        'created_at': datetime(...)
      },
      ...
    ]
    """
    if filename is None:
        os.makedirs("reports", exist_ok=True)
        filename = f"reports/user_actions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    os.makedirs(os.path.dirname(filename), exist_ok=True)

    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(["user_id", "telegram_id", "username", "action", "details", "created_at"])
        for a in actions:
            writer.writerow([
                a.get("user_id"),
                a.get("telegram_id"),
                a.get("username") or "",
                a.get("action") or "",
                a.get("details") or "",
                a.get("created_at").strftime("%Y-%m-%d %H:%M:%S")
                if a.get("created_at") else ""
            ])

    return filename


def export_user_actions_to_excel(actions, filename=None):
    """
    Экспорт действий пользователей в Excel файл.
    
    actions — список словарей с информацией о действиях пользователей
    """
    if filename is None:
        os.makedirs("reports", exist_ok=True)
        filename = f"reports/user_actions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    os.makedirs(os.path.dirname(filename), exist_ok=True)

    # Создаём рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Логи действий"

    # Стили
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Заголовки
    headers = ["ID пользователя", "Telegram ID", "Имя пользователя", "Действие", "Детали", "Дата/Время"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Данные
    for row_num, action in enumerate(actions, 2):
        cells_data = [
            action.get("user_id"),
            action.get("telegram_id"),
            action.get("username") or "",
            action.get("action") or "",
            action.get("details") or "",
            action.get("created_at").strftime("%Y-%m-%d %H:%M:%S") if action.get("created_at") else ""
        ]
        
        for col_num, value in enumerate(cells_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = center_align if col_num < 4 else left_align
            cell.border = border

    # Автоширина колонок
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20

    wb.save(filename)
    logger.info(f"Экспорт логов в Excel: {filename}")
    return filename


def export_schedule_to_excel(schedule_data, group_name=None, filename=None):
    """
    Экспорт расписания в Excel файл.
    
    schedule_data — список словарей:
    [
      {
        'group_number': 'БПИ-24',
        'lesson_date': date,
        'lesson_number': 1,
        'start_time': '09:00',
        'end_time': '10:30',
        'subject_name': 'Программирование',
        'subject_type': 'lecture',
        'teacher_fio': 'Иванов И.И.',
        'building_name': 'Корпус 1',
        'room_number': '101',
        'notes': '...'
      },
      ...
    ]
    """
    if filename is None:
        os.makedirs("reports", exist_ok=True)
        group_text = group_name if group_name else "all"
        filename = f"reports/schedule_{group_text}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    os.makedirs(os.path.dirname(filename), exist_ok=True)

    # Создаём рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"

    # Стили
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Заголовок
    title = f"Расписание занятий" + (f" - {group_name}" if group_name else "")
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = center_align

    ws.append([])  # пустая строка

    # Заголовки колонок
    headers = ["Группа", "Дата", "День", "Пара", "Время", "Предмет", "Тип", "Преподаватель", "Аудитория"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Данные
    for row_num, item in enumerate(schedule_data, 4):
        lesson_date = item.get('lesson_date')
        
        # Преобразуем дату если нужно
        if hasattr(lesson_date, 'strftime'):
            date_str = lesson_date.strftime("%Y-%m-%d")
            day_name = get_day_name(lesson_date.weekday())
        else:
            date_str = str(lesson_date)
            day_name = "?"

        cells_data = [
            item.get('group_number', ''),
            date_str,
            day_name,
            item.get('lesson_number', ''),
            f"{item.get('start_time', '')}-{item.get('end_time', '')}",
            item.get('subject_name', ''),
            item.get('subject_type', ''),
            item.get('teacher_fio', ''),
            f"{item.get('room_number', '')} ({item.get('building_name', '')})" if item.get('room_number') else ""
        ]
        
        for col_num, value in enumerate(cells_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = center_align if col_num <= 5 else left_align
            cell.border = border

    # Автоширина колонок
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20

    wb.save(filename)
    logger.info(f"Экспорт расписания в Excel: {filename}")
    return filename


def create_schedule_import_template(filename=None):
    """
    Создает шаблон Excel для импорта расписания.
    Структура: Группа | Дата | Пара | Начало | Конец | Предмет | Тип | Преподаватель | Аудитория
    """
    if filename is None:
        os.makedirs("reports", exist_ok=True)
        filename = f"reports/schedule_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    os.makedirs(os.path.dirname(filename), exist_ok=True)

    # Создаём рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"

    # Стили
    header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    example_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    example_font = Font(italic=True, color="999999", size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Заголовок
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = "ШАБЛОН ИМПОРТА РАСПИСАНИЯ"
    title_cell.font = Font(bold=True, size=14, color="FF6B6B")
    title_cell.alignment = center_align

    # Заголовки колонок
    headers = ["Группа*", "Дата*", "Пара*", "Начало*", "Конец*", "Предмет*", "Тип*", "Преподаватель", "Аудитория"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Описание и примеры
    examples = [
        ["БПИ-24", "2026-02-10", "1", "09:00", "10:30", "Программирование", "lecture", "Иванов И.И.", "101"],
        ["БПИ-24", "2026-02-10", "2", "10:45", "12:15", "Алгоритмы", "practice", "Петров П.П.", "102"],
        ["БПИ-25", "2026-02-10", "1", "09:00", "10:30", "Математика", "lecture", "Сидоров С.С.", "201"],
    ]

    for row_num, example in enumerate(examples, 4):
        for col_num, value in enumerate(example, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.fill = example_fill
            cell.font = example_font
            cell.alignment = center_align if col_num in [2, 3, 6, 7] else left_align
            cell.border = border

    # Легенда
    legend_row = 8
    ws.merge_cells(f'A{legend_row}:I{legend_row}')
    legend_cell = ws[f'A{legend_row}']
    legend_cell.value = "* - обязательные поля | Дата формат: YYYY-MM-DD | Время формат: HH:MM | Типы: lecture, practice, lab"
    legend_cell.font = Font(italic=True, size=9, color="666666")
    legend_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Путанные типы пар
    types_row = 10
    ws.merge_cells(f'A{types_row}:I{types_row}')
    types_cell = ws[f'A{types_row}']
    types_cell.value = "Доступные типы: lecture (лекция), practice (практика), lab (лабораторная)"
    types_cell.font = Font(italic=True, size=9, color="4472C4")

    # Размеры колонок
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15

    wb.save(filename)
    logger.info(f"Создан шаблон импорта расписания: {filename}")
    return filename


def import_schedule_from_excel(filepath, db_manager=None):
    """
    Импорт расписания из Excel файла.
    
    Возвращает:
    {
        'success': True/False,
        'message': 'Сообщение об ошибке или успехе',
        'added': количество добавленных записей,
        'errors': список ошибок
    }
    """
    errors = []
    added_count = 0
    result = {
        'success': False,
        'message': '',
        'added': 0,
        'errors': []
    }

    try:
        # Загружаем файл
        wb = load_workbook(filepath)
        ws = wb.active

        schedule_records = []

        # Пропускаем первые 3 строки (заголовок и описание)
        for row_num in range(4, ws.max_row + 1):
            row_data = []
            for col_num in range(1, 10):  # 9 колонок
                cell = ws.cell(row=row_num, column=col_num)
                row_data.append(cell.value)

            # Пропускаем пустые строки
            if not any(row_data[:5]):  # Если первые 5 обязательных полей пусты
                continue

            try:
                group = str(row_data[0]).strip() if row_data[0] else None
                date_str = str(row_data[1]).strip() if row_data[1] else None
                lesson_num = int(row_data[2]) if row_data[2] else None
                start_time = str(row_data[3]).strip() if row_data[3] else None
                end_time = str(row_data[4]).strip() if row_data[4] else None
                subject = str(row_data[5]).strip() if row_data[5] else None
                lesson_type = str(row_data[6]).strip() if row_data[6] else None
                teacher = str(row_data[7]).strip() if row_data[7] else None
                room = str(row_data[8]).strip() if row_data[8] else None

                # Валидация
                if not all([group, date_str, lesson_num, start_time, end_time, subject, lesson_type]):
                    errors.append(f"Строка {row_num}: Не все обязательные поля заполнены")
                    continue

                # Валидация типа пары
                if lesson_type not in ['lecture', 'practice', 'lab']:
                    errors.append(f"Строка {row_num}: Неизвестный тип пары '{lesson_type}'")
                    continue

                schedule_records.append({
                    'group': group,
                    'date': date_str,
                    'lesson_number': lesson_num,
                    'start_time': start_time,
                    'end_time': end_time,
                    'subject': subject,
                    'subject_type': lesson_type,
                    'teacher': teacher if teacher else None,
                    'room': room if room else None
                })

            except Exception as e:
                errors.append(f"Строка {row_num}: {str(e)}")

        if not schedule_records:
            result['message'] = "Нет данных для импорта"
            result['errors'] = errors
            return result

        # Добавляем в БД если передан db_manager
        if db_manager:
            for record in schedule_records:
                try:
                    db_manager.add_schedule_from_import(
                        group_number=record['group'],
                        lesson_date=record['date'],
                        lesson_number=record['lesson_number'],
                        start_time=record['start_time'],
                        end_time=record['end_time'],
                        subject_name=record['subject'],
                        subject_type=record['subject_type'],
                        teacher_fio=record['teacher'],
                        room_number=record['room']
                    )
                    added_count += 1
                except Exception as e:
                    errors.append(f"Ошибка импорта: Группа {record['group']}, дата {record['date']}: {str(e)}")

        result['success'] = True
        result['message'] = f"Успешно импортировано {added_count} записей"
        result['added'] = added_count
        result['errors'] = errors

        logger.info(f"Импорт расписания завершен: {added_count} записей, {len(errors)} ошибок")

    except Exception as e:
        result['message'] = f"Ошибка при открытии файла: {str(e)}"
        logger.error(f"Ошибка импорта расписания: {e}")

    return result


def get_day_name(weekday_num):
    """Преобразование номера дня недели в название"""
    days = {
        0: "Понедельник",
        1: "Вторник",
        2: "Среда",
        3: "Четверг",
        4: "Пятница",
        5: "Суббота",
        6: "Воскресенье"
    }
    return days.get(weekday_num, "?")

